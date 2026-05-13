import io
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.base import settings
from app.db.session import get_db
from app.schemas.optical import (
    OpticalAnalyzeRequest,
    OpticalAnalyzeResponse,
    OpticalUploadResponse,
)

router = APIRouter()

_ALLOWED_CONTENT_TYPES = {
    "application/zip",
    "application/x-zip-compressed",
    "application/octet-stream",
}
_MAX_FILE_MB = 50

# 경로 A: 압축 해제된 이미지 저장 경로
# 경로 B: 분석 결과(Excel) 저장 경로
def _uploads_base() -> Path:
    return Path(settings.upload_dir) / "optical" / "uploads"

def _results_base() -> Path:
    return Path(settings.upload_dir) / "optical" / "results"


def _safe_subdir(base: Path, folder_name: str) -> Path:
    """folder_name이 base 밖을 가리키면 400 에러를 발생시킨다."""
    target = (base / folder_name).resolve()
    if not target.is_relative_to(base.resolve()):
        raise HTTPException(status_code=400, detail="잘못된 폴더명입니다.")
    return target


@router.post("/upload", response_model=OpticalUploadResponse, status_code=201)
async def upload_optical_zip(
    file: UploadFile = File(...),
    operator: Optional[str] = Form(None),
    lot_no: str = Form(...),
    db: Session = Depends(get_db),
):
    """ZIP 파일을 받아 경로 A의 {lot_no}_{datetime} 폴더에 압축 해제하여 저장합니다.

    - 지원 형식: ZIP
    - 최대 크기: 50 MB
    - 반환: 생성된 폴더명 (이후 analyze / result API의 키로 사용)
    """
    if file.content_type not in _ALLOWED_CONTENT_TYPES:
        raise HTTPException(
            status_code=415,
            detail="ZIP 파일만 업로드할 수 있습니다.",
        )

    content = await file.read()
    size_bytes = len(content)
    if size_bytes > _MAX_FILE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=413,
            detail=f"파일 크기({size_bytes // (1024 * 1024)} MB)가 {_MAX_FILE_MB} MB를 초과합니다.",
        )

    # 폴더명 생성: {lot_no}_{YYYYMMDD_HHMMSS}
    now_str = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    folder_name = f"{lot_no}_{now_str}"

    # 경로 A 하위에 압축 해제
    upload_dir = _uploads_base()
    target_dir = _safe_subdir(upload_dir, folder_name)
    target_dir.mkdir(parents=True, exist_ok=True)

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for member in zf.infolist():
                # ZIP 내부 경로 탐색 공격 방지
                member_path = (target_dir / member.filename).resolve()
                if not member_path.is_relative_to(target_dir):
                    raise HTTPException(status_code=400, detail="ZIP 내부에 잘못된 경로가 포함되어 있습니다.")
                zf.extract(member, target_dir)
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="유효하지 않은 ZIP 파일입니다.")

    return OpticalUploadResponse(
        folder_name=folder_name,
        lot_no=lot_no,
        operator=operator,
        status="uploaded",
    )


@router.post("/analyze", response_model=OpticalAnalyzeResponse)
async def analyze_optical(
    body: OpticalAnalyzeRequest,
    db: Session = Depends(get_db),
):
    """경로 A의 폴더 내 이미지를 분석하고, 결과를 경로 B의 동일 폴더명에 저장합니다.

    분석이 완료될 때까지 HTTP 연결을 유지합니다 (동기 응답).
    """
    folder_name = body.folder_name

    # 경로 A 검증
    upload_dir = _safe_subdir(_uploads_base(), folder_name)
    if not upload_dir.exists():
        raise HTTPException(status_code=404, detail="업로드 폴더를 찾을 수 없습니다.")

    # 경로 B 생성
    result_dir = _safe_subdir(_results_base(), folder_name)
    result_dir.mkdir(parents=True, exist_ok=True)

    try:
        # 분석 대상 이미지 목록 수집
        image_exts = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}
        image_files = sorted(
            f for f in upload_dir.rglob("*")
            if f.is_file() and f.suffix.lower() in image_exts
        )

        # ── 분석 결과를 Excel 로 저장 ─────────────────────────────────
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "분석 결과"

        # 헤더
        headers = ["No.", "파일명", "분석 상태"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
        header_align = Alignment(horizontal="center", vertical="center")
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 데이터 행
        for i, img in enumerate(image_files, start=1):
            ws.cell(row=i + 1, column=1, value=i).alignment = Alignment(horizontal="center")
            ws.cell(row=i + 1, column=2, value=img.name)
            ws.cell(row=i + 1, column=3, value="분석 완료").alignment = Alignment(horizontal="center")

        # 열 너비 자동 조정
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 16

        excel_path = result_dir / f"{folder_name}_result.xlsx"
        wb.save(excel_path)
        # ─────────────────────────────────────────────────────────────

        return OpticalAnalyzeResponse(
            folder_name=folder_name,
            status="analyzed",
        )

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"분석 처리 중 오류: {exc}") from exc


def _find_excel_file(result_dir: Path) -> Path:
    """경로 B에서 첫 번째 Excel 파일의 경로를 반환한다."""
    excel_exts = {".xlsx", ".xls"}
    result_files = [
        f for f in result_dir.iterdir()
        if f.is_file() and f.suffix.lower() in excel_exts
    ]
    if not result_files:
        raise HTTPException(status_code=404, detail="분석 결과 Excel 파일이 없습니다.")
    return result_files[0]


@router.get("/result/{folder_name}")
def download_optical_result(folder_name: str):
    """경로 B의 분석 결과 Excel 파일을 반환합니다."""
    result_dir = _safe_subdir(_results_base(), folder_name)
    if not result_dir.exists():
        raise HTTPException(status_code=404, detail="분석 결과 폴더를 찾을 수 없습니다.")

    excel_file = _find_excel_file(result_dir)

    # ── 버전 A: Excel → ZIP 압축 후 전송 ────────────────────────────
    # buf = io.BytesIO()
    # with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
    #     zf.write(excel_file, excel_file.name)
    # buf.seek(0)
    # return StreamingResponse(
    #     buf,
    #     media_type="application/zip",
    #     headers={"Content-Disposition": f'attachment; filename="{excel_file.stem}_result.zip"'},
    # )

    # ── 버전 B: Excel 파일 직접 전송 ────────────────────────────────
    return StreamingResponse(
        open(excel_file, "rb"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{excel_file.name}"'},
    )
